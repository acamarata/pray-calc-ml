"""
Process India Twilight Photometer data (Pune/Kolhapur).

Station info (from paper: Mane, Pratibha B. — Earth and Space Science):
  Location: Kolhapur, Maharashtra, India
  Coordinates: ~16.7°N, 74.2°E (Kolhapur) or Pune 18.52°N, 73.86°E
  Elevation: ~570m (Kolhapur) / 565m (Pune)
  Timezone: IST = UTC+5:30

The twilight photometer measures atmospheric airglow/scattering intensity
(arbitrary linear units) at 10-second cadence during civil and nautical
twilight. Higher values = brighter sky (more solar illumination).

Each Excel file contains multiple observation sessions in groups of 3
columns: [datetime_string | intensity | None].

The datetime strings are local IST times formatted as "YYYY.MM.DD HH:MM:SS".

Morning sessions (05:XX–07:XX IST) cover depression ~16° → ~4° (Fajr range).
Evening sessions (18:XX–19:XX IST) cover depression ~2° → ~10° (partial Isha).

Processing:
- Parse each session column group
- Convert IST to UTC (subtract 5h30m)
- Compute solar depression via PyEphem
- Find readings within depression 10°-18° window
- Find inflection point (max |d(intensity)/dt|)
- Output one row per twilight event

Note on coordinates: The paper covers Kolhapur primarily. Using
16.7°N, 74.2°E as the station location.
"""

import os
import datetime
import ephem
import pandas as pd
import openpyxl

# Kolhapur, Maharashtra, India
LAT = 16.7
LON = 74.2
ELEV = 570.0
UTC_OFFSET = 5.5  # IST = UTC+5:30

INDIA_DIR = "/Volumes/X9/Sites/acamarata/pray-calc-ml/data/raw/india_twilight_photometer"
OUT_CSV = "/Volumes/X9/Sites/acamarata/pray-calc-ml/data/raw/raw_sightings/india_twilight_photometer.csv"
SOURCE = "India_Twilight_Photometer_Zenodo5541367"

DEP_MIN = 10.0
DEP_MAX = 18.0  # photometer sessions don't often go past ~18°


def compute_solar_depression(utc_dt: datetime.datetime) -> float:
    obs = ephem.Observer()
    obs.lat = str(LAT)
    obs.lon = str(LON)
    obs.elev = ELEV
    obs.pressure = 0
    obs.epoch = ephem.J2000
    obs.date = utc_dt.strftime("%Y/%m/%d %H:%M:%S")
    sun = ephem.Sun()
    sun.compute(obs)
    alt_deg = float(sun.alt) * 180.0 / ephem.pi
    return -alt_deg


def parse_session(ws, col_start: int) -> pd.DataFrame:
    """
    Parse one session (col_start = datetime col, col_start+1 = intensity).
    Returns DataFrame with utc_dt, intensity columns.
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        dt_val = row[col_start] if col_start < len(row) else None
        int_val = row[col_start + 1] if col_start + 1 < len(row) else None
        if dt_val is None or int_val is None:
            continue
        if not isinstance(dt_val, str):
            continue
        try:
            local_dt = datetime.datetime.strptime(dt_val.strip(), "%Y.%m.%d %H:%M:%S")
        except ValueError:
            continue
        try:
            intensity = float(int_val)
        except (TypeError, ValueError):
            continue
        # Convert IST to UTC: subtract 5h30m
        utc_dt = local_dt - datetime.timedelta(hours=5, minutes=30)
        rows.append({"utc_dt": utc_dt, "local_dt": local_dt, "intensity": intensity})

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).sort_values("utc_dt").reset_index(drop=True)
    return df


def find_twilight_event(df: pd.DataFrame, prayer: str) -> dict | None:
    """
    Given a session DataFrame, find the twilight inflection point
    within the depression window and return event dict.
    """
    if len(df) < 4:
        return None

    # Compute solar depression for each reading
    depressions = [compute_solar_depression(r["utc_dt"]) for _, r in df.iterrows()]
    df = df.copy()
    df["depression"] = depressions

    # Filter to window
    mask = (df["depression"] >= DEP_MIN) & (df["depression"] <= DEP_MAX)
    window = df[mask].copy()
    if len(window) < 3:
        return None

    window = window.sort_values("utc_dt").reset_index(drop=True)
    window["dintensity"] = window["intensity"].diff()
    window["ddep"] = window["depression"].diff()

    # For morning (fajr): depression decreasing, intensity increasing
    # For evening (isha): depression increasing, intensity decreasing
    if prayer == "fajr":
        segment = window[window["ddep"] < -0.01].copy()
    else:
        segment = window[window["ddep"] > 0.01].copy()

    if len(segment) < 3:
        # Fall back to all window data
        segment = window.copy()

    segment["abs_dintensity"] = segment["dintensity"].abs()
    if segment.empty or segment["abs_dintensity"].isna().all():
        return None

    peak_idx = segment["abs_dintensity"].idxmax()
    peak_row = segment.loc[peak_idx]

    utc_dt = peak_row["utc_dt"]
    dep = peak_row["depression"]
    local_dt = peak_row["local_dt"]

    return {
        "prayer": prayer,
        "date_local": local_dt.strftime("%Y-%m-%d"),
        "time_local": local_dt.strftime("%H:%M:%S"),
        "utc_offset": UTC_OFFSET,
        "lat": LAT,
        "lng": LON,
        "elevation_m": ELEV,
        "source": SOURCE,
        "notes": (
            f"photometer_intensity={peak_row['intensity']:.3f},"
            f"solar_dep={dep:.2f}deg,inflection_method"
        ),
    }


def classify_session(df: pd.DataFrame) -> str:
    """Determine if a session is morning (fajr) or evening (isha) by local time."""
    if df.empty:
        return "unknown"
    hour = df.iloc[0]["local_dt"].hour
    # Morning: before 12:00 local, Evening: after 12:00 local
    return "fajr" if hour < 12 else "isha"


def main():
    all_events = []

    xlsx_files = sorted(
        f for f in os.listdir(INDIA_DIR) if f.endswith(".xlsx")
    )
    print(f"Processing {len(xlsx_files)} Excel files...")

    for fname in xlsx_files:
        fpath = os.path.join(INDIA_DIR, fname)
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active

        n_cols = ws.max_column
        # Sessions in groups of 3 columns (datetime, value, None gap)
        session_count = 0

        for col_start in range(0, n_cols - 1, 3):
            df = parse_session(ws, col_start)
            if df.empty:
                continue

            prayer = classify_session(df)
            if prayer == "unknown":
                continue

            event = find_twilight_event(df, prayer)
            if event:
                all_events.append(event)
                session_count += 1

        print(f"  {fname}: {session_count} events extracted")
        wb.close()

    if not all_events:
        print("No events found.")
        return

    out_df = pd.DataFrame(all_events, columns=[
        "prayer", "date_local", "time_local", "utc_offset",
        "lat", "lng", "elevation_m", "source", "notes"
    ])
    out_df = out_df.sort_values(["date_local", "prayer"]).reset_index(drop=True)
    out_df.to_csv(OUT_CSV, index=False)
    print(f"\nWrote {len(out_df)} rows to {OUT_CSV}")
    print(out_df["prayer"].value_counts().to_string())


if __name__ == "__main__":
    main()
